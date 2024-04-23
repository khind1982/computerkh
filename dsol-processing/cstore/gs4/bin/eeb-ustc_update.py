#!/usr/bin/env python2.7

# Script to update eeb handover records with USTC enriched data

import os
import re
import sys
import shutil

sys.path.append('/packages/dsol/opt/lib/python2.7/site-packages')
sys.path.append('/packages/dsol/platform/lib')

import lxml.etree as ET

from openpyxl import load_workbook
from openpyxl.utils import coordinate_from_string, column_index_from_string
from xml.sax.saxutils import unescape

from commonUtils.fileUtils import locate
from commonUtils.textUtils import replace_entityrefs
from leviathanUtils import eeboutorder


parser = ET.XMLParser(remove_blank_text=True)

try:
    inlog = sys.argv[1]
    indir = sys.argv[2]
    outdir = sys.argv[3]
except IndexError:
    print "Usage: eeb-ustc_update.py [input file] [input dir] [output dir]\n"
    exit(1)

if indir.replace('/', '') == outdir.replace('/', ''):
    print("Input directory cannot be the same as the output directory\n"
    "Usage: eeb-ustc_update.py [inputfile] [input directory] [output directory]\n")
    exit(1)

if not os.path.exists(outdir):
    os.makedirs(outdir)

print 'Loading workbook...'

wb = load_workbook(filename = inlog)
sheet = wb['To-EEB-Map']

rows_to_update = {}
updates = {}

def presenceTestSingle(element):
    if element is not None and element.text is not None:
        return element.text
    else:
        return None

def hexToDecimal(string):
    match = re.search('&#(x[0-9A-z]+);', string)
    if match:
        print match.group(1)
        dec = int('0%s' % match.group(1), 16)
        return string.replace(match.group(1), str(dec))
    else:
        return string


def errorFile(outdir, error):
    with open('%s/errorlog.txt' % outdir, 'w+') as out:
        out.write(error + '\n')


rownum = 0
headers = {}

pqidlist = [os.path.basename(i).replace('.xml', '') for i in locate('*.xml', indir)]
loclist = [i for i in locate('*.xml', indir)]

titledict = {}

print 'Working out pqid column and compiling title dictionary...'

for column in sheet.iter_cols():
    if column[0].value == 'pqid':
        sheetpqids = [cell.value for cell in column]
        pqidcol = column_index_from_string(column[0].column) - 1
    if column[0].value == 'title':
        for cell in column:
            if cell.value != '' and cell.value != None:
                pqid = sheet.cell(row=cell.row, column=pqidcol + 1).value
                if type(cell.value) != long and type(cell.value) != float:
                    titledict[pqid] = str(cell.value.encode('utf-8'))
                else:
                    titledict[pqid] = str(cell.value)

comp = [i for i in loclist if os.path.basename(i).replace('.xml', '') not in sheetpqids]

for unenriched in comp:
    shutil.copyfile(unenriched, '%s/%s' % (outdir, os.path.basename(unenriched)))

print 'Enriching files...'

for row in sheet.iter_rows():
    rowdict = {}
    pqid = row[pqidcol].value
    rownum += 1
    if row[pqidcol].value in pqidlist:
        infile = [f for f in loclist if row[pqidcol].value in f]
        # print "Adding enrichment data to file %s..." % (infile[0])
        if os.path.basename(infile[0]).replace('.xml', '') not in sheetpqids:
            print infile[0], ' not in master enrichment'
        for cell in row:
            if cell.value != None:
                if cell.value != '':
                    coll = cell.column
                    collrow1 = sheet['%s1' % coll].value
                    if collrow1 != 'pqid':
                        if collrow1 not in rowdict.keys():
                            # print collrow1
                            rowdict[collrow1] = []
                        if type(cell.value) != long and type(cell.value) != float:
                            rowdict[collrow1].append(str(cell.value.encode('utf-8')))
                        else:
                            rowdict[collrow1].append(str(cell.value))

        root = ET.parse(infile[0], parser).getroot()
        rec_search = root.find('.//rec_search')

        #this updates fields in the XML with fields from the spreadsheet
        # print "Processing data from row %s in spreadsheet..." % (rownum)
        for key in rowdict.keys():
            if len(root.findall('.//%s' % key)) != 0:
                for ele in root.findall('.//%s' % key):
                    ele.getparent().remove(ele)
            for i in rowdict[key]:
                i = i.decode('utf-8').encode('ascii', 'xmlcharrefreplace'
                    ).replace('&', '&amp;')
                # i = re.sub('&([^a|^#])', '&amp;\1', i)
                if '/' in key:
                    try:
                        split_key = key.split('/')[0]
                        parent = rec_search.find('.//%s' % split_key)
                        try:
                            ET.SubElement(parent, key.split('/')[1]).text = i.strip()
                        except ValueError:
                            print('Row %s column %s cell value %s is not '
                                'Unicode or ASCII compatible' % (rownum, key, i.strip()))
                    except TypeError:
                        print "No element '%s' in %s. Cannot continue." % (split_key, pqid)
                        sys.exit(13)
                else:
                    try:
                        ET.SubElement(rec_search, key).text = i.strip()
                    except ValueError:
                        print('Row %s column %s cell value %s is not '
                            'Unicode or ASCII compatible' % (rownum, key, i.strip()))

        #this updates the link titles

        links = root.findall('.//link')

        for link in links:
            linkid = link.find('.//linkid')
            linktitle = link.find('.//linktitle')
            if linkid.text in titledict.keys():
                try:
                    linktitle.text = titledict[linkid.text].decode('utf-8')
                except ValueError:
                    print titledict[linkid.text]
                    raise

            # sys.stdout.write('Total rows: %s Rows processed: %s\r' % (sheet.max_row, rownum))
            # sys.stdout.flush()

        outf = '%s/%s' % (outdir, os.path.basename(infile[0]))

        root = eeboutorder(root)
        with open(outf, 'w') as out:
            out.write(ET.tostring(root, pretty_print=True).decode('utf-8'
                ).replace('&amp;amp;', '&amp;'
                ).replace('&amp;#', '&#'))
