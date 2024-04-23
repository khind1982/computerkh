#!/usr/bin/env python2.7
# -*- coding: utf-8 -*-

# Script to output log and titlelists

import os
import re
import sys

sys.path.append(os.path.join(os.path.dirname(__file__), '..', 'lib'))
sys.path.append(os.path.join(os.path.dirname(__file__), '..', 'libdata/eeb'))
sys.path.append('/packages/dsol/opt/lib/python2.7/site-packages')

import lxml.etree as ET
from openpyxl import load_workbook, Workbook
import xlrd
import xlwt
from xlutils.copy import copy

from StringIO import StringIO
from commonUtils.fileUtils import locate, buildLut


# To make a new workbook and save it
# wb = Workbook()

# sheet = wb.active

# wb.save('/home/cmoore/prob-solve/test.xlsx')

wb = load_workbook('/home/cmoore/prob-solve/test.xlsx')

sheet = wb.active

wb.save('/home/cmoore/prob-solve/test.xlsx')

updates_book = load_workbook('/home/cmoore/prob-solve/eeb-logger/eeb_incoming_log-test.xlsx')
updates_sheet = updates_book['Log']
updates_book.save('/home/cmoore/prob-solve/eeb_incoming_log-test.xlsx')