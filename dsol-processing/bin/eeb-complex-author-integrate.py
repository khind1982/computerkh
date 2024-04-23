#!/usr/local/bin/python2.7

# Create a global dictionary keyed on author whose values are filenames in a list
# For each author name in the component sheet, create a new object which separates
# the names into its component parts.  Call the new class 'authComponents'.
# Desired variables in 'authComponents': corrected_form, last_name, first_name,
# middle_name, birth_date, death_date, title, profession, corporate
# The components will be returned as a tuple
# Match the corrected_form with a key in auth_file_dict - if match then get
# the list of files.  Iterate over the list of files.  Open each file and find
# the <author_corrected> element in the <author_main> and <author_other> elements.
# Match the text of the element with the first item in the tuple.
# If the items match, insert the additional elements in the new author elements.
# /rec/rec_search/author_main|author_other/author_name
# /rec/rec_search/author_main|author_other/author_corrected
# /rec/rec_search/author_main|author_other/author_uninverted
# /rec/rec_search/author_main|author_other/DoNotNormalize
# /rec/rec_search/author_main|author_other/last_name
# /rec/rec_search/author_main|author_other/first_name
# /rec/rec_search/author_main|author_other/middle_name
# /rec/rec_search/author_main|author_other/birth_date
# /rec/rec_search/author_main|author_other/death_date
# /rec/rec_search/author_main|author_other/person_title
# /rec/rec_search/author_main|author_other/professional_title
# /rec/rec_search/author_main|author_other/author_corporate

# Desired new behaviour:
#   Create the auth_file_dict but use the filenames as keys
#   The values of the auth file dict should be a list of component objects
#   Iterate over the data in any input directory.
#   If a file is found with the same filename as one in the auth_file_dict:
#   Find the author in the file with the same author_corrected field
#   Update the file with the new elements


import sys
import os

from collections import defaultdict

sys.path.append("/packages/dsol/opt/lib/python2.7/site-packages")

import openpyxl
import difflib
import codecs
import lxml.etree as et


class authfileCollector(object):

    def __init__(self):
        self.auth_file_dict = defaultdict(list)

    def collect_auths_files(self, fname_sheet):
        for row in fname_sheet.iter_rows(min_row=2):
            self.auth_file_dict[row[1].value].append(row[0].value)

    def compare_authors(self, components):
        for file, authlist in self.auth_file_dict.items():
            # print auth, components[0]
            for auth in authlist:
                if auth == components[0]:
                    return authlist


class authComponents(object):

    def __init__(self):
        self.corrected_form = ''
        self.last_name = ''
        self.first_name = ''
        self.middle_name = ''
        self.birth_date = ''
        self.death_date = ''
        self.title = ''
        self.profession = ''
        self.corporate = ''

    def create_auth_components(self, row):
        self.corrected_form = row[0].value
        self.last_name = row[1].value
        self.first_name = row[2].value
        self.middle_name = row[3].value
        self.birth_date = row[4].value
        self.death_date = row[5].value
        self.title = row[6].value
        self.profession = row[7].value
        self.corporate = row[8].value
        return self


def main(args):
    try:
        auth_names_wksheet = args[0]
        modify_files_dir = args[1]
        outdir = args[2]
    except IndexError:
        print("Usage: [spreadsheet of author names] [path to directory of files to modify] [path to output directory]")
        exit (1)

    if not os.path.exists(outdir):
        os.makedirs(outdir)

    wb = openpyxl.load_workbook(auth_names_wksheet)
    fname_sheet = wb["names-files"]
    component_sheet = wb["names-components"]

    parser = et.XMLParser(remove_blank_text=True)

    # Create the defaultdict that holds the files associated with each author
    authfilecollect = authfileCollector()
    authfilecollect.collect_auths_files(fname_sheet)

    seen = 0

    # Build a dictionary consisting of author names to authcomponent objects
    authcomponentsdict = {}
    for row in component_sheet.iter_rows(min_row=2):
        authcomponent = authComponents()
        seen += 1
        print("Processing row %s in spreadsheet..." % seen)
        components = authcomponent.create_auth_components(row)
        authcomponentsdict[components.corrected_form] = components


    rec = 0
    with open('%s/errors.txt' % (outdir), 'w') as errorfile:
        for f in walk_dirs(modify_files_dir):
            rec += 1
            print("Processing record %s, (%s)" % (f, rec))
            xmldata = et.parse(f, parser)
            docid = os.path.basename(f)
            collection = xmldata.xpath('.//unit')[0].text
            library = docid.split('-')[1]
            outpath = os.path.join(outdir, collection, library)
            if not os.path.exists(outpath):
                os.makedirs(outpath)
            if docid in authfilecollect.auth_file_dict.keys() or '%s.xml' % (docid) in authfilecollect.auth_file_dict.keys():
                authors = authfilecollect.auth_file_dict[docid]
                if len(authors) > 1:
                    errorfile.write("Note: File %s contains more than one author with complex author names.  Check all authors updated." % (docid))
                for author in authors:
                    match = check_author_corrected_elements(xmldata, author)
                # If there's no match, add to the error log
                    if match != False:
                        try:
                            components = authcomponentsdict[author]
                        except KeyError:
                            errorfile.write("Error: Target author name %s in file %s cannot be found or does not match name in the component sheet.\n" % (author, docid))
                            continue
                        author_element_to_modify = match
                        modified_author = modify_auth_element(author_element_to_modify, components)
                        with open(os.path.join(outpath, docid), 'w') as outf:
                            outf.write(et.tostring(modified_author.getroottree(), pretty_print=True).decode())
                    else:
                        errorfile.write("Error: Target author name %s in file %s cannot be found or does not match name in the component sheet.\n" % (author, docid))
            else:
                with open(os.path.join(outpath, docid), 'w') as outf:
                    outf.write(et.tostring(xmldata, pretty_print=True).decode())



def check_author_corrected_elements(xmldata, author):
    for author_corrected in xmldata.xpath('.//author_main/author_corrected|.//author_other/author_corrected'):
        if author == et.tostring(author_corrected).decode().split('<author_corrected>')[1].split('</author_corrected>')[0]:
            return author_corrected.getparent()
    return False


def modify_auth_element(auth_elem, components):
    if auth_elem.xpath('./last_name'):
        for elem in auth_elem.xpath('./last_name'):
            et.strip_elements(auth_elem, 'last_name')
    if auth_elem.xpath('./first_name'):
        for elem in auth_elem.xpath('./first_name'):
            et.strip_elements(auth_elem, 'first_name')
    if auth_elem.xpath('./middle_name'):
        for elem in auth_elem.xpath('./middle_name'):
            et.strip_elements(auth_elem, 'middle_name')
    if auth_elem.xpath('./birth_date'):
        for elem in auth_elem.xpath('./birth_date'):
            et.strip_elements(auth_elem, 'birth_date')
    if auth_elem.xpath('./death_date'):
        for elem in auth_elem.xpath('./death_date'):
            et.strip_elements(auth_elem, 'death_date')
    if auth_elem.xpath('./person_title'):
        for elem in auth_elem.xpath('./person_title'):
            et.strip_elements(auth_elem, 'person_title')
    if auth_elem.xpath('./professional_title'):
        for elem in auth_elem.xpath('./professional_title'):
            et.strip_elements(auth_elem, 'professional_title')
    if auth_elem.xpath('./author_corporate'):
        for elem in auth_elem.xpath('./author_corporate'):
            et.strip_elements(auth_elem, 'author_corporate')
    if auth_elem.xpath('./DoNotNormalize'):
        for elem in auth_elem.xpath('./DoNotNormalize'):
            et.strip_elements(auth_elem, 'DoNotNormalize')
    normalize = et.SubElement(auth_elem, 'DoNotNormalize')
    normalize.text = "Y"
    if components.last_name:
        last_name_element = et.SubElement(auth_elem, 'last_name')
        last_name_element.text = components.last_name
    if components.first_name:
        first_name_element = et.SubElement(auth_elem, 'first_name')
        first_name_element.text = components.first_name
    if components.middle_name:
        middle_name_element = et.SubElement(auth_elem, 'middle_name')
        middle_name_element.text = components.middle_name
    if components.birth_date:
        birth_date_element = et.SubElement(auth_elem, 'birth_date')
        birth_date_element.text = str(components.birth_date)
    if components.death_date:
        death_date_element = et.SubElement(auth_elem, 'death_date')
        death_date_element.text = str(components.death_date)
    if components.title:
        title_element = et.SubElement(auth_elem, 'person_title')
        title_element.text = components.title
    if components.profession:
        profession_element = et.SubElement(auth_elem, 'professional_title')
        profession_element.text = components.profession
    if components.corporate:
        corporate_author_element = et.SubElement(auth_elem, 'author_corporate')
        corporate_author_element.text = components.corporate
    return auth_elem


def walk_dirs(indir):
    for root, dirs, files in os.walk(indir):
        for f in files:
            if f.endswith('.xml'):
                yield os.path.join(root, f)


if __name__ == "__main__":
    main(sys.argv[1:])
