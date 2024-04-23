#!/usr/bin/env python2.7
# -*- coding: utf-8 -*-

'''This script checks that all packages in a certain directory have been logged'''

import os
import glob
import re
import sys

sys.path.append(os.path.join(os.path.dirname(__file__), '..', 'lib'))
sys.path.append('/packages/dsol/opt/lib/python2.7/site-packages')

import lxml.etree as ET
import shutil
from commonUtils.fileUtils import locate, buildLut
from openpyxl import load_workbook
from openpyxl.utils import coordinate_from_string, column_index_from_string


try:
    indir = sys.argv[1]
    depth = sys.argv[2]
except IndexError:
    print "Usage: eeb-check_packages_logged.py [input directory] [bookid depth]\n"
    exit(1)

locs = []
pqids = []

print 'Loading workbook...\n'

wb = load_workbook(filename = '/dc/eurobo/documentation/tracking/eeb_incoming_log.xlsx')
sheet = wb['Log']

for column in sheet.iter_cols():
    if column[0].value == 'Location':
        for i in column:
            split = i.value.split('/')
            new = '/'.join(split[0:-1])
            locs.append(new)

if depth == '2':
    filesdepth = glob.glob('%s/*/*' % indir)
elif depth == '1':
    filesdepth = glob.glob('%s/*' % indir)
elif depth == '3':
    filesdepth = glob.glob('%s/*/*/*' % indir)

print 'Checking that packages are logged...\n'
dirsDepth2 = filter(lambda f: os.path.isdir(f), filesdepth)

# Checks that pqid directories appear in the log
for i in dirsDepth2:
    if i not in locs and i.endswith('-000') != True:
        print 'Warning! Directory %s not logged!' % (i)
