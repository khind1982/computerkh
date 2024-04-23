#!/usr/bin/env python2.7

# Script to update eeb handover records with USTC enriched data

import os
import re
import sys

sys.path.append('/packages/dsol/opt/lib/python2.7/site-packages')
sys.path.append('/packages/dsol/platform/lib')

import lxml.etree as ET
from commonUtils.fileUtils import locate
from openpyxl import load_workbook
from openpyxl.utils import coordinate_from_string, column_index_from_string

parser = ET.XMLParser(remove_blank_text=True)


try:
    inlog = sys.argv[1]
    outlog = sys.argv[2]
    indir = sys.argv[3]
except IndexError:
    print "Usage: eeb-log_update.py [input file] [output file] [input dir]\n"
    exit(1)

wb = load_workbook(filename = inlog)
sheet = wb.active

rows_to_update = {}
updates = {}

def presenceTestSingle(element):
    if element is not None and element.text is not None:
        return element.text
    else:
        return None

print 'Building update dictionary\n'

for f in locate('*.xml', indir):

    sys.stdout.write('File:%s\r' % f)
    sys.stdout.flush()

    update_data = {}    
    root = ET.parse(f, parser).getroot()

    update_data['Status'] = 'Published: %s' % root.find('.//unit').text
    update_data['Title'] =  presenceTestSingle(root.find('.//title'))
    update_data['Author'] = presenceTestSingle(root.find('.//author_main/author_corrected'))
    update_data['Pubdate'] = presenceTestSingle(root.find('.//displaydate'))
    update_data['Imprint'] = presenceTestSingle(root.find('.//imprint'))
    update_data['Shelfmark'] = presenceTestSingle(root.find('.//shelfmark'))
    update_data['Country'] = presenceTestSingle(root.find('.//country_of_publication'))
    update_data['Language'] = '|'.join([i.text for i in root.findall('.//language')])
    update_data['Subject'] = '|'.join([i.text for i in root.findall('.//subject')])

    for i in enumerate(root.findall('.//classification1')):
        update_data['USTC Classification %s' % str(i[0] + 1)] = i[1].text

    updates[root.find('.//itemid').text] = update_data

print '\nApplying updates to SS in memory'

rownum = 0
headers = {}

for column in sheet.iter_cols():
    if column[0].value == 'PQID':
        pqidcol = column_index_from_string(column[0].column) - 1

for row in sheet.iter_rows():
    rownum += 1
    sys.stdout.write('Rows seen: %s\r' % rownum)
    sys.stdout.flush()
    if row[pqidcol].value in updates.keys():
        for cell in row:
            coll = cell.column
            collrow1 = sheet['%s1' % coll].value
            if collrow1 in updates[row[pqidcol].value].keys():
                if updates[row[pqidcol].value][collrow1] is not None:
                    cell.value = updates[row[pqidcol].value][collrow1]

print 'Writing updated workbook'

wb.save(outlog)